import os
import warnings

os.environ["CUDA_VISIBLE_DEVICES"] = '0'
warnings.filterwarnings('ignore')

from networks.PraNet_Res2Net_TTA import PraNet
from config import *
from utils.metrics import calculate_metrics
import numpy as np
import argparse
import time
import torch
import openpyxl
import itertools
from torch.utils.data import DataLoader
from dataloaders.POLYP_dataloader import POLYP_dataset
from dataloaders.convert_csv_to_list import convert_labeled_list
from torch.autograd import Variable

torch.set_num_threads(1)


class RunTTA:
    def __init__(self, config):
        self.load_model = os.path.join(config.path_save_model, str(config.Source_Dataset))
        self.log_path = os.path.join(config.path_save_log, 'TTA')

        if not os.path.exists(self.log_path):
            os.makedirs(self.log_path)

        target_test_csv = []
        for target in config.Target_Dataset:
            target_test_csv.append(target + '_train.csv')
            target_test_csv.append(target + '_test.csv')
        ts_img_list, ts_label_list = convert_labeled_list(config.dataset_root, target_test_csv)

        target_test_dataset = POLYP_dataset(config.dataset_root, ts_img_list, ts_label_list,
                                            config.image_size)
        self.target_test_loader = DataLoader(dataset=target_test_dataset,
                                             batch_size=config.batch_size,
                                             shuffle=False,
                                             pin_memory=True,
                                             drop_last=False,
                                             num_workers=config.num_workers)

        self.in_ch = config.in_ch
        self.out_ch = config.out_ch
        self.image_size = config.image_size
        self.model_type = config.model_type
        self.device = config.device

        for arg, value in vars(config).items():
            print(f"{arg}: {value}")
        print('***' * 20)

        self.build_model()
        self.print_network()

    def build_model(self):
        self.model = PraNet().to(self.device)
        checkpoint = torch.load(os.path.join(self.load_model, 'pretrain-PraNet.pth'))
        self.model.load_state_dict(checkpoint, strict=True)

    def print_network(self):
        num_params = 0
        for p in self.model.parameters():
            num_params += p.numel()
        print("The number of total parameters: {}".format(num_params))

    def run(self):
        metrics_test = [[], [], []]
        metric_dict = ['Dice', 'Enhanced_Align', 'Structure_Measure']

        self.model.eval()
        for batch, data in enumerate(self.target_test_loader):
            x, y, path = data
            x, y = Variable(x).to(self.device), Variable(y).to(self.device)

            pred_logit = self.model(x)
            seg_output = torch.sigmoid(pred_logit)

            metrics = calculate_metrics(seg_output.detach().cpu(), y.detach().cpu())
            for i in range(len(metrics)):
                assert isinstance(metrics[i], list), "The metrics value is not list type."
                metrics_test[i] += metrics[i]

        test_metrics_y = np.mean(metrics_test, axis=1)
        print_test_metric_mean = {}
        for i in range(len(test_metrics_y)):
            print_test_metric_mean[metric_dict[i]] = test_metrics_y[i]
        print("Test Metrics Mean: ", print_test_metric_mean)

        test_metrics_y = np.std(metrics_test, axis=1)
        print_test_metric_std = {}
        for i in range(len(test_metrics_y)):
            print_test_metric_std[metric_dict[i]] = test_metrics_y[i]
        print("Test Metrics Std: ", print_test_metric_std)
        return print_test_metric_mean


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--model_type', type=str, default='PraNet', help='PraNet')
    parser.add_argument('--in_ch', type=int, default=3)
    parser.add_argument('--out_ch', type=int, default=1)
    parser.add_argument('--image_size', type=int, default=352)
    parser.add_argument('--num_workers', type=int, default=8)
    parser.add_argument('--batch_size', type=int, default=1)

    parser.add_argument('--Source_Dataset', type=str,
                        help='BKAI/CVC-ClinicDB/ETIS-LaribPolypDB/Kvasir-SEG')
    parser.add_argument('--Target_Dataset', type=list,
                        help='BKAI/CVC-ClinicDB/ETIS-LaribPolypDB/Kvasir-SEG')

    parser.add_argument('--path_save_model', type=str, default='./models/')
    parser.add_argument('--path_save_log', type=str, default='./logs/')
    parser.add_argument('--dataset_root', type=str, default='/media/userdisk3/yhfu/Datasets/Polyp')

    if torch.cuda.is_available():
        parser.add_argument('--device', type=str, default='cuda:0')
    else:
        parser.add_argument('--device', type=str, default='cpu')
    config = parser.parse_args()

    start_time = time.time()
    all_datasets = ['BKAI', 'CVC-ClinicDB', 'ETIS-LaribPolypDB', 'Kvasir-SEG']
    combinations = list(itertools.permutations(all_datasets, 2))
    work_book = openpyxl.Workbook()

    sheet_names = ['Dice', 'Enhanced_Align', 'Structure_Measure']
    for idx, sheet_name in enumerate(sheet_names):
        work_book.create_sheet(title=sheet_name, index=idx)

    for sheet_name in sheet_names:
        for i, dataset in enumerate(all_datasets, start=2):
            work_book[sheet_name].cell(row=1, column=i, value=dataset)
            work_book[sheet_name].cell(row=i, column=1, value=dataset)

    for source_dataset, target_dataset in combinations:
        config.Source_Dataset = source_dataset
        config.Target_Dataset = [target_dataset]
        TTA = RunTTA(config)

        print('=====' * 10, '\n')
        print('Load Pretrained Source Dataset:', config.Source_Dataset)
        print('Target Dataset:', config.Target_Dataset)
        metric = TTA.run()
        print('Metric:', metric)
        print('=====' * 10, '\n')

        row = all_datasets.index(source_dataset) + 2
        col = all_datasets.index(target_dataset) + 2
        for sheet_name in sheet_names:
            work_book[sheet_name].cell(row=row, column=col, value=round(metric[sheet_name], 3))
    work_book.save(os.path.join(config.path_save_log, 'TTA', 'DyNo.xlsx'))

    end_time = time.time()
    duration = (end_time - start_time)
    print(f"duration：{duration:.3f} seconds")
